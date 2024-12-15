from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QTableView, QLineEdit, QToolBar, QStatusBar, QSplitter, QHeaderView, 
    QMessageBox, QMenu, QPushButton, QFileDialog, QSizePolicy,QListWidget
)
from .edit_window_ui import Ui_EditWindow
import json
import re
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, QVariant, pyqtSignal, QEvent
from PyQt6.QtGui import QAction, QKeySequence, QBrush, QColor, QUndoStack, QUndoCommand, QActionGroup, QFont
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import configparser
import os
from utils.validators import format_datetime, is_date
from zoneinfo import ZoneInfo
import pytz

class ShiftDatesCommand(QUndoCommand):
    def __init__(self, model, days, original_data):
        super().__init__(f"Shift Dates by {days} days")
        self.model = model
        self.days = days
        self.original_data = original_data #Corrected attribute name

    def undo(self):
        self.model.set_dates(self.original_data) #Corrected attribute name

    def redo(self):
        self.model.shift_dates(self.days)


class FindReplaceCommand(QUndoCommand):
    def __init__(self, model, view, find_text, replace_text): # Add view parameter
        super().__init__("Find and Replace")
        self.model = model
        self.view = view # Store the view
        self.find_text = find_text
        self.replace_text = replace_text
        self.original_data = self.model.get_data_frame().copy()


    def undo(self):
        self.model.set_data(self.original_data)

    def redo(self):
            for row in range(self.model.rowCount()):
                for col in range(self.model.columnCount()):
                    index = self.model.index(row, col)
                    cell_value = str(self.model.data(index, Qt.ItemDataRole.DisplayRole))
                    if self.find_text.lower() in cell_value.lower():
                        self.model.setData(index, self.replace_text, Qt.ItemDataRole.EditRole) # Removed the updated variable and return statement




class EditCommand(QUndoCommand):
    def __init__(self, model, index, old_value, new_value):
        super().__init__("Edit Cell")
        self.model = model
        self.row = index.row()
        self.column = index.column()
        self.old_value = old_value
        self.new_value = new_value

    def undo(self):
        self.model._data_frame.iloc[self.row, self.column] = self.old_value
        self.model.dataChanged.emit(self.model.index(self.row, self.column), self.model.index(self.row, self.column), [Qt.ItemDataRole.EditRole])

    def redo(self):
        self.model._data_frame.iloc[self.row, self.column] = self.new_value
        self.model.dataChanged.emit(self.model.index(self.row, self.column), self.model.index(self.row, self.column), [Qt.ItemDataRole.EditRole])
        
class IncrementEpisodeNumberCommand(QUndoCommand):
    def __init__(self, model, table_view): # Add table_view as a parameter
        super().__init__("Increment Episode Number")
        self.model = model
        self.table_view = table_view # Store table_view for later use
        self.original_data = self.model.get_data_frame().copy() #store a copy of the original DataFrame


    def undo(self):
        self.model.set_data(self.original_data) #Restore from the copy of the original DataFrame

    def redo(self):
        self.model.increment_episode_number_filtered(self.table_view) #Call filtered function directly


class DataFrameModel(QAbstractTableModel):
    def __init__(self, data_frame: pd.DataFrame, undo_stack):
        super().__init__()
        self._data_frame = data_frame.copy()
        self._original_data_frame = data_frame.copy()
        self.undo_stack = undo_stack

    def rowCount(self, parent=None):
        return self._data_frame.shape[0]

    def columnCount(self, parent=None):
        return self._data_frame.shape[1]

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return QVariant()
        value = self._data_frame.iloc[index.row(), index.column()]
        if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
            return str(value)
        elif role == Qt.ItemDataRole.BackgroundRole:
            if not self.is_cell_valid(index):
                return QBrush(QColor('#ffcccc'))
        return QVariant()
    
    
    def set_data(self, data_frame):
        self._data_frame = data_frame.copy()
        self.dataChanged.emit(self.index(0, 0), self.index(self.rowCount() - 1, self.columnCount() - 1), [Qt.ItemDataRole.EditRole])
        self.layoutChanged.emit() # Emit layoutChanged as well for a complete refresh

    def is_cell_valid(self, index):
        value = self._data_frame.iloc[index.row(), index.column()]
        column_name = self._data_frame.columns[index.column()]
        required_columns = ['DATE', 'START TIME', 'NAZIV EMISIJE']
        if column_name in required_columns:
            if pd.isna(value) or str(value).strip() == '':
                return False
            if column_name == 'DATE':
                return is_date(str(value))
        return True
    
    def set_dates(self, data):
        self._data_frame = data.copy()  #Directly set DataFrame
        self.dataChanged.emit(self.index(0, 0), self.index(self.rowCount() - 1, self.columnCount() - 1), [Qt.ItemDataRole.EditRole])
        self.recalculate_stop_times()

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return QVariant()
        if orientation == Qt.Orientation.Horizontal:
            column_name = self._data_frame.columns[section]
            if column_name == 'DATE':
                return "DATUM"
            elif column_name == 'START TIME':
                return "POČETAK"
            elif column_name == 'NAZIV EMISIJE':
                return "NAZIV EMISIJE"
            elif column_name == 'CATEGORY':
                return "KATEGORIJA"
            elif column_name == 'EPISODE NUMBER':
                return "NUMERACIJA"
            elif column_name == 'P/R':
                return "REPRIZA"
            elif column_name == 'OPIS emisije':
                return "OPIS EMISIJE"
            else:
                return column_name  # Return original name for other columns
        else:
            return str(section)

    def flags(self, index):
        return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if index.isValid() and role == Qt.ItemDataRole.EditRole:
            row = index.row()
            col = index.column()
            column_name = self._data_frame.columns[index.column()]

            if column_name == 'START TIME':
                if isinstance(value, str):
                    match = re.fullmatch(r'\d{4}', value) # Check for 4-digit input (HHMM)
                    if match:
                        value = value[:2] + ":" + value[2:] #Insert colon

            old_value = self._data_frame.iloc[row, col]
            if str(old_value) != str(value):
                command = EditCommand(self, index, old_value, value)
                self.undo_stack.push(command)
                self._data_frame.iloc[row, col] = value
                self.dataChanged.emit(index, index, [role])
                return True  #Removed call to on_data_changed
        return False

    def insertRows(self, position, rows=1, parent=QModelIndex()):
        self.beginInsertRows(QModelIndex(), position, position + rows - 1)
        empty_row = pd.DataFrame([[""] * self.columnCount()], columns=self._data_frame.columns)
        self._data_frame = pd.concat([self._data_frame.iloc[:position], empty_row, self._data_frame.iloc[position:]]).reset_index(drop=True)
        self.endInsertRows()
        return True

    def removeRows(self, position, rows=1, parent=QModelIndex()):
        removed_rows = self._data_frame.iloc[position:position + rows] #added
        self.beginRemoveRows(QModelIndex(), position, position + rows - 1)
        self._data_frame = self._data_frame.drop(self._data_frame.index[position:position + rows]).reset_index(drop=True)
        self.endRemoveRows()
        command = RemoveRowsCommand(self, position, rows, removed_rows) #added
        self.undo_stack.push(command) #added
        return True

    def get_data_frame(self):
        return self._data_frame.copy()

    def reset_data(self):
        self._data_frame = self._original_data_frame.copy()
        self.layoutChanged.emit()

    def has_unsaved_changes(self):
        return not self._data_frame.equals(self._original_data_frame)
    
    def shift_dates(self, days):
        date_column_index = self._data_frame.columns.get_loc('DATE')
        
        for i in range(self.rowCount()):
            index = self.index(i, date_column_index)
            current_date_str = self.data(index, Qt.ItemDataRole.DisplayRole)
            if is_date(current_date_str):
                current_date = datetime.strptime(current_date_str, '%d.%m.%Y.').date()
                new_date = current_date + timedelta(days=days)
                new_date_str = new_date.strftime('%d.%m.%Y.')
                self._data_frame.iloc[i, date_column_index] = new_date_str #Directly modify DataFrame

        self.recalculate_stop_times()
        self.dataChanged.emit(self.index(0, 0), self.index(self.rowCount() - 1, self.columnCount() - 1), [Qt.ItemDataRole.EditRole])


    def recalculate_stop_times(self):
        timezone = ZoneInfo("Europe/Zagreb")
        try:
            self._data_frame['start'] = pd.to_datetime(self._data_frame['DATE'].astype(str) + ' ' + self._data_frame['START TIME'].astype(str), format='%d.%m.%Y. %H:%M', errors='raise').dt.tz_localize(timezone, ambiguous='NaT')
            self._data_frame['stop'] = self._data_frame['start'].shift(-1)

            last_program_start = self._data_frame['start'].iloc[-1]
            next_day = last_program_start + timedelta(days=1)
            stop_dt = next_day.replace(hour=7, minute=0, second=0, microsecond=0)
            self._data_frame['stop'] = self._data_frame['stop'].fillna(stop_dt)

            self.dataChanged.emit(self.index(0, 0), self.index(self.rowCount() - 1, self.columnCount() - 1), [Qt.ItemDataRole.EditRole])

        except ValueError as e:
            QMessageBox.warning(self, "Greška pri konverziji datuma/vremena", str(e))
        except pytz.exceptions.AmbiguousTimeError as e:
            QMessageBox.warning(self, "Nejednoznačno vrijeme", f"Došlo je do nejednoznačnosti u vremenu: {e}. Provjerite svoje podatke.")
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Neočekivana pogreška: {e}")
            
    def increment_episode_number_filtered(self, table_view):
            ep_num_col = self._data_frame.columns.get_loc('EPISODE NUMBER')
            for row in range(self.rowCount()):
                if not table_view.isRowHidden(row):
                    index = self.index(row, ep_num_col)
                    current_value = self.data(index, Qt.ItemDataRole.DisplayRole)

                    try:
                        if current_value is not None and str(current_value).strip() != "":
                            if isinstance(current_value, (int, float)):
                                new_value = int(current_value) + 1
                            elif current_value.isdigit():
                                new_value = int(current_value) + 1
                            elif '-' in current_value:
                                parts = current_value.split('-')
                                if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                                    new_value = f"{int(parts[0]) + 1}-{int(parts[1]) + 1}"
                                else:
                                    continue
                            else:
                                continue

                            # Directly modify the DataFrame
                            self._data_frame.iloc[row, ep_num_col] = new_value

                            # Emit dataChanged signal to update the view
                            self.dataChanged.emit(index, index, [Qt.ItemDataRole.EditRole])

                    except (ValueError, TypeError) as e:
                        print(f"Error incrementing episode number in row {row}: {e}")


   
class RemoveRowsCommand(QUndoCommand):
    def __init__(self, model, position, rows, removed_rows):
        super().__init__("Remove Rows")
        self.model = model
        self.position = position
        self.rows = rows
        self.removed_rows = removed_rows

    def undo(self):
        self.model.beginInsertRows(QModelIndex(), self.position, self.position + self.rows - 1)
        self.model._data_frame = pd.concat([self.model._data_frame.iloc[:self.position], self.removed_rows, self.model._data_frame.iloc[self.position:]]).reset_index(drop=True)
        self.model.endInsertRows()

    def redo(self):
        self.model.beginRemoveRows(QModelIndex(), self.position, self.position + self.rows - 1)
        self.model._data_frame = self.model._data_frame.drop(self.model._data_frame.index[self.position:self.position + self.rows]).reset_index(drop=True)
        self.model.endRemoveRows()


class EditWindow(QDialog):
    data_saved = pyqtSignal()

    def __init__(self, display_df, internal_df, excel_file_path, excel_save_dir, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"UREDI RASPORED PROGRAMA - {os.path.basename(excel_file_path)}") # Modified title
        self.setGeometry(100, 100, 900, 700)
        self.setObjectName("EditWindow") 

        self.display_df = display_df
        self.internal_df = internal_df
        self.excel_file_path = excel_file_path
        self.excel_save_dir = excel_save_dir

        self.undo_stack = QUndoStack(self)
        self.unsaved_changes = False
        
        # Enable maximizing
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint)
        

        # Initialize UI components
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        self.setLayout(main_layout)

        # Toolbar
        self.toolbar = QToolBar("Toolbar", self)
        main_layout.addWidget(self.toolbar)


        # Search and Replace Fields
        search_replace_layout = QHBoxLayout()
        main_layout.addLayout(search_replace_layout)

        self.find_label = QLabel("   Traži:")
        search_replace_layout.addWidget(self.find_label)

        self.find_field = QLineEdit(self)
        search_replace_layout.addWidget(self.find_field)

        self.replace_label = QLabel("   Zamijeni sa:")
        search_replace_layout.addWidget(self.replace_label)

        self.replace_field = QLineEdit(self)
        search_replace_layout.addWidget(self.replace_field)

        self.replace_button = QPushButton("Zamijeni", self)
        self.replace_button.clicked.connect(self.find_and_replace)
        search_replace_layout.addWidget(self.replace_button)

        # Table View with proper layout
        splitter = QSplitter(Qt.Orientation.Vertical)
        main_layout.addWidget(splitter)
        

        # CREATE table_view HERE
        self.table_view = QTableView(splitter)
        self.table_model = DataFrameModel(self.display_df, self.undo_stack)
        self.table_view.setModel(self.table_model)


        self.table_view.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.customContextMenuRequested.connect(self.open_context_menu)
        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.installEventFilter(self)
        self.table_view.selectionModel().selectionChanged.connect(self.on_selection_changed) #Connect to selectionChanged

        splitter.addWidget(self.table_view)

        self.status_bar = QStatusBar()  # Create the status bar
        self.status_bar.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed) #added
        splitter.addWidget(self.status_bar)
        # Connect signals for UI logic
        self.init_ui_logic()
        


    def init_ui_logic(self):
        # Create action groups
        save_group = QActionGroup(self)
        edit_group = QActionGroup(self)
        undo_group = QActionGroup(self)
        close_group = QActionGroup(self)
        self.find_field.textChanged.connect(self.search) 
        self.table_model.dataChanged.connect(self.on_data_changed)

        # Add actions to groups and toolbar, with separators
        save_action = self.toolbar.addAction("Spremi")
        save_action.triggered.connect(self.save_changes)  # You should have the connection here, not higher up in init_ui
        save_group.addAction(save_action)
        self.toolbar.addSeparator() # Separator after edit group
        
        save_as_action = self.toolbar.addAction("Spremi kao")
        save_as_action.triggered.connect(self.save_as_workbook) # Same: connect triggered in init_ui_logic, not in init_ui 
        save_group.addAction(save_as_action)
        self.toolbar.addSeparator()  # Separator after Save group
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator() 
        

        add_action = self.toolbar.addAction("Dodaj red")
        add_action.triggered.connect(self.add_row)  # Moved connection here
        edit_group.addAction(add_action)
        self.toolbar.addSeparator() # Separator after edit group

        delete_action = self.toolbar.addAction("Obriši red")
        delete_action.triggered.connect(self.delete_row) # Moved connection here
        edit_group.addAction(delete_action)
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator() 

        undo_action = self.toolbar.addAction("Poništi")
        undo_action.triggered.connect(self.undo_stack.undo) # Moved connection here
        undo_group.addAction(undo_action)
        self.toolbar.addSeparator() # Separator after edit group
        

        redo_action = self.toolbar.addAction("Ponovi")
        redo_action.triggered.connect(self.undo_stack.redo)  # Moved connection here
        undo_group.addAction(redo_action)
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator()
        self.toolbar.addSeparator() 
        self.toolbar.addSeparator() 
        
        increment_episode_number_action = QAction("Numeracija +1", self)
        increment_episode_number_action.triggered.connect(self.increment_episode_number)
        self.toolbar.addAction(increment_episode_number_action)
        self.toolbar.addSeparator()
        self.toolbar.addSeparator() # Separator after edit group
        self.toolbar.addSeparator()
        self.toolbar.addSeparator() 
        self.toolbar.addSeparator() 

        add_7_days_action = QAction("Datum +7", self)
        add_7_days_action.triggered.connect(lambda: self.shift_dates(7))
        self.toolbar.addAction(add_7_days_action)
        self.toolbar.addSeparator()      
        self.toolbar.addSeparator()  
        self.toolbar.addSeparator()
        self.toolbar.addSeparator()
        self.toolbar.addSeparator() 
        self.toolbar.addSeparator() 

        
        close_action = self.toolbar.addAction("Zatvori")
        close_action.triggered.connect(self.close)
        close_group.addAction(close_action)
        
    def increment_episode_number(self):
        command = IncrementEpisodeNumberCommand(self.table_model, self.table_view)
        self.undo_stack.push(command)  # Pass the table_view to the command
        
        
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def drop_event(self, event):
        mimeData = event.mimeData()
        if mimeData.hasUrls():
            for url in mimeData.urls():
                filePath = str(url.toLocalFile())
                if filePath.lower().endswith(('.xls', '.xlsx')):
                    self.load_excel(filePath)

        else:
            event.ignore()

    def save_column_widths(self):
        config = configparser.ConfigParser()
        config_path = os.path.join(self.excel_save_dir, 'config.ini')

        # Ensure config.ini exists; create it if it doesn't.
        os.makedirs(self.excel_save_dir, exist_ok=True) #create directory if not exists
        if not os.path.exists(config_path):
            with open(config_path, 'w') as configfile:
                config.write(configfile)

        try:
            config.read(config_path)  # Read existing config

            # Ensure the 'ColumnWidths' section exists, create it if it doesn't.
            if 'ColumnWidths' not in config:
                config['ColumnWidths'] = {}

            widths = [self.table_view.columnWidth(i) for i in range(self.table_model.columnCount())]
            config['ColumnWidths']['widths'] = json.dumps(widths)  # Use JSON for better handling

            with open(config_path, 'w') as configfile:
                config.write(configfile)  # Write the updated config

        except Exception as e:
            print(f"Error saving column widths: {e}")


    def load_column_widths(self):
        config = configparser.ConfigParser()
        config_path = os.path.join(self.excel_save_dir, "config.ini")
        if os.path.exists(config_path):
            try:
                config.read(config_path)
                if 'ColumnWidths' in config and 'widths' in config['ColumnWidths']:
                    widths_str = config['ColumnWidths']['widths']
                    try:
                        widths = json.loads(widths_str) # Use json.loads for deserialization
                        if len(widths) == self.table_model.columnCount():
                            for i, width in enumerate(widths):
                                self.table_view.setColumnWidth(i, width)
                    except json.JSONDecodeError:
                        print("Error parsing column widths from config file.")
            except Exception as e:
                print(f"Error loading column widths: {e}")
    

    def save_as_workbook(self):
        if not self.validate_data():
            QMessageBox.warning(self, "Upozorenje", "Podaci nisu valjani. Ispravite ih prije spremanja.")
            return

        new_file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Spremi kao",
            self.excel_save_dir,
            "Excel Files (*.xlsx)"
        )
        
        if new_file_name:
            if not new_file_name.endswith('.xlsx'):
                new_file_name += '.xlsx'
            
            try:
                self.save_workbook(new_file_name)
                self.excel_file_path = new_file_name
                self.unsaved_changes = False
                self.data_saved.emit()
                
            except Exception as e:
                self.status_bar.showMessage("Greška pri spremanju datoteke", 5000)
                QMessageBox.critical(self, "Greška", f"Došlo je do greške pri spremanju:\n{e}")
        self.status_bar.showMessage("Workbook je spremljen.", 5000)
        self.unsaved_changes = False # Set to False after successful save
        self.data_saved.emit() # Emit the signal



    def save_changes(self):
        if not self.validate_data():
            QMessageBox.warning(self, "Upozorenje", "Podaci nisu valjani. Ispravite ih prije spremanja.")
            return

        try:
            self.save_workbook(self.excel_file_path)
            self.unsaved_changes = False
            self.status_bar.showMessage("Promjene su uspješno spremljene.", 5000)
            self.data_saved.emit()
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Došlo je do greške prilikom spremanja:\n{str(e)}")
            
                          
    def save_workbook(self, save_path):
        updated_df = self.table_model.get_data_frame()
        for col in updated_df.columns:
            if pd.api.types.is_datetime64_any_dtype(updated_df[col]):
                updated_df[col] = updated_df[col].dt.tz_localize(None)
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        for column in range(len(updated_df.columns)):
            sheet.cell(row=1, column=column+1).value = updated_df.columns[column]
            sheet.cell(row=1, column=column+1).font = Font(bold=True)
            sheet.cell(row=1, column=column+1).alignment = Alignment(horizontal="left")

        # Write data
        for row in range(len(updated_df)):
            for column in range(len(updated_df.columns)):
                sheet.cell(row=row+2, column=column+1).value = updated_df.iloc[row, column]

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Apply table style
        table_style = openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table = openpyxl.worksheet.table.Table(
            ref=f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}",
            displayName="DiadoraTV",
            tableStyleInfo=table_style,
        )
        sheet.add_table(table)
        
        # Add banded rows
        light_gray_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid') # Light gray
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if row[0].row % 2 == 0:  # Apply to every other row starting from the second row (data rows)
                for cell in row:
                    cell.fill = light_gray_fill

        # Add borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

        # Special formatting for "OPIS emisije"
        opis_column_index = updated_df.columns.get_loc("OPIS emisije") + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=opis_column_index)
            cell.alignment = Alignment(horizontal="left", wrap_text=True)

        workbook.save(save_path)
        self.status_bar.showMessage("Promjene su uspješno spremljene.", 5000)
        QMessageBox.information(self, "Uspjeh", f"Promjene su uspješno spremljene u {save_path}.")
        self.unsaved_changes = False
        self.table_model._original_data_frame = updated_df.copy()
        self.undo_stack.clear()
        self.data_saved.emit()
        
    def on_data_changed(self, topLeft, bottomRight, roles):
        """Handles data changes in the table view."""
        if self.table_model.has_unsaved_changes():  # Check if data is actually different
            self.unsaved_changes = True
            self.status_bar.showMessage("Postoje nespremljene promjene.", 2000)  # Only show the message if there are unsaved changes

        if not self.validate_data(): #add validation check
            self.status_bar.showMessage("Podaci nisu valjani. Postoje nespremljene promjene.", 5000) #add a status bar message


        self.table_view.viewport().update()  # Refresh the view regardless
        
    def on_selection_changed(self, selected, deselected):
        indexes = self.table_view.selectedIndexes()
        if indexes:
            row = indexes[0].row()
            column = indexes[0].column()

            # Highlight row number
            row_header_text = self.table_view.verticalHeader().model().headerData(row, Qt.Orientation.Vertical, Qt.ItemDataRole.DisplayRole)
            if row_header_text:
                brush = QBrush(QColor("lightgreen"))
                self.table_view.verticalHeader().model().setData(self.table_view.verticalHeader().model().index(row, 0), brush, Qt.ItemDataRole.ForegroundRole)

            # Highlight column header
            col_header_text = self.table_view.horizontalHeader().model().headerData(column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
            if col_header_text:
                brush = QBrush(QColor("lightgreen"))
                self.table_view.horizontalHeader().model().setData(self.table_view.horizontalHeader().model().index(0, column), brush, Qt.ItemDataRole.ForegroundRole)


    def validate_data(self):
        df = self.table_model.get_data_frame()
        timezone = ZoneInfo("Europe/Zagreb")

        try:
            df['start'] = pd.to_datetime(df['DATE'].astype(str) + ' ' + df['START TIME'].astype(str), format='%d.%m.%Y. %H:%M', errors='raise').dt.tz_localize(timezone, ambiguous='NaT')
            df['stop'] = df['start'].shift(-1)

            last_row_index = len(df) - 1
            last_start = df['start'].iloc[last_row_index]
            df.loc[last_row_index, 'stop'] = last_start + timedelta(days=1, hours=7)

            required_fields = ['DATE', 'START TIME', 'NAZIV EMISIJE']
            if df[required_fields].isnull().values.any():
                QMessageBox.warning(self, "Validation Error", "Nedostaju obvezna polja.")
                return False

            for i in range(len(df) - 1):
                if df['start'].iloc[i] >= df['stop'].iloc[i + 1]:
                    #Simplified warning message
                    QMessageBox.warning(self, "Vremensko preklapanje!", f"Preklapanje vremena! Provjerite retke {i + 0} i {i + 1}")
                    return False

            return True

        except (ValueError, TypeError) as e:
            QMessageBox.warning(self, "Greška pri konverziji datuma/vremena", str(e))
            return False
        except Exception as e:
            QMessageBox.critical(self, "Greška", f"Neočekivana pogreška: {e}")
            return False
    
    def shift_dates(self, days):
        try:
            original_data = self.table_model.get_data_frame().copy()  # Store the entire DataFrame
            command = ShiftDatesCommand(self.table_model, days, original_data) # Pass the entire DataFrame
            self.undo_stack.push(command)

          #  self.table_model.shift_dates(days)  # Perform the date shift
            self.unsaved_changes = True
            self.status_bar.showMessage(f"Datumi u stupcu 'DATUM' pomaknuti za {days} dana.", 5000)
            self.table_view.viewport().update()

        except Exception as e:
            QMessageBox.warning(self, "Greška", f"Došlo je do greške prilikom promjene datuma: {e}")

                                 
    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if index.isValid() and role == Qt.ItemDataRole.EditRole:
            row = index.row()
            col = index.column()
            old_value = self._data_frame.iloc[row, col]
            if str(old_value) != str(value):
                command = EditCommand(self, index, old_value, value)
                self.undo_stack.push(command)
                self._data_frame.iloc[row, col] = value
                self.dataChanged.emit(index, index, [role])
                self.on_data_changed(index, index, [role]) # Added
                return True
        return False

    
    def add_row(self):
        current_row = self.table_view.currentIndex().row()
        if current_row == -1:
            current_row = self.table_model.rowCount()
        self.table_model.insertRows(current_row)
        self.unsaved_changes = True

    def delete_row(self):
        current_row = self.table_view.currentIndex().row()
        if current_row >= 0:
            reply = QMessageBox.question(
                self, "Potvrda brisanja", "Jeste li sigurni da želite obrisati odabrani red?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.table_model.removeRows(current_row)
                self.unsaved_changes = True
        else:
            QMessageBox.warning(self, "Upozorenje", "Niste odabrali red za brisanje.")

    def search(self, text):
        # Store initial state (scroll position and selection)
        initial_scroll = self.table_view.verticalScrollBar().value()
        initial_selection = self.table_view.currentIndex()

        # Hide/show rows based on search
        for row in range(self.table_model.rowCount()):
            match = any(
                text.lower() in str(self.table_model.data(self.table_model.index(row, col), Qt.ItemDataRole.DisplayRole)).lower()
                for col in range(self.table_model.columnCount())
            )
            self.table_view.setRowHidden(row, not match)

        if not text:  # Search cleared
            # Find the first visible row
            first_visible_row = -1
            for i in range(self.table_model.rowCount()):
                if not self.table_view.isRowHidden(i):
                    first_visible_row = i
                    break

            # Force a model refresh and then scroll
            self.table_model.layoutChanged.emit()
            self.table_view.reset()  #Crucial for re-rendering after layout changes

            if first_visible_row != -1:
                index = self.table_model.index(first_visible_row, 0)

                # Calculate the center position
                row_height = self.table_view.rowHeight(first_visible_row)
                viewport_height = self.table_view.viewport().height()
                visible_rows = viewport_height // row_height
                center_row = first_visible_row - visible_rows // 2

                # Bound the center row (handle cases where the center row is outside table bounds).
                center_row = max(0, min(center_row, self.table_model.rowCount() - 1))


                center_index = self.table_model.index(center_row, 0)  # Create index from the calculated center row

                self.table_view.scrollTo(center_index, QTableView.ScrollHint.EnsureVisible)


            else:
                # If no visible rows, restore the initial scroll position
                self.table_view.verticalScrollBar().setValue(initial_scroll)

            # Restore initial selection if valid and the row is still visible.
            if initial_selection.isValid() and not self.table_view.isRowHidden(initial_selection.row()):
                self.table_view.setCurrentIndex(initial_selection)

            self.table_view.viewport().update()


    def find_and_replace(self):
        find_text = self.find_field.text()
        replace_text = self.replace_field.text()

        if not find_text:
            QMessageBox.warning(self, "Upozorenje", "Unesite tekst za pretraživanje.")
            return

        original_data = self.table_model.get_data_frame().copy()  # Save original data for undo
        command = FindReplaceCommand(self.table_model, find_text, replace_text, original_data)
        self.undo_stack.push(command)  # Push the command onto the undo stack

        updated = False
        for row in range(self.table_model.rowCount()):
            for col in range(self.table_model.columnCount()):
                index = self.table_model.index(row, col)
                cell_value = str(self.table_model.data(index, Qt.ItemDataRole.DisplayRole))
                if find_text.lower() in cell_value.lower():  # Case-insensitive substring search
                    if self.table_model.setData(index, replace_text, Qt.ItemDataRole.EditRole): # Crucial change for correct setData usage
                        updated = True

        if updated:
            self.status_bar.showMessage("Zamjena izvršena.", 5000)
            self.unsaved_changes = True # Correctly sets the unsaved_changes flag
        else:
            QMessageBox.information(self, "Obavijest", "Traženi tekst nije pronađen.")
            
    

    def open_context_menu(self, position):
        menu = QMenu()

        add_action = QAction("Dodaj red", self)
        add_action.triggered.connect(self.add_row)
        menu.addAction(add_action)

        delete_action = QAction("Obriši red", self)
        delete_action.triggered.connect(self.delete_row)
        menu.addAction(delete_action)

        search_action = QAction("Pretraži", self)
        search_action.triggered.connect(self.search_selected_cell)
        menu.addAction(search_action)

        menu.addSeparator()

        undo_action = QAction("Poništi", self)
        undo_action.triggered.connect(self.undo_stack.undo)
        menu.addAction(undo_action)

        redo_action = QAction("Ponovi", self)
        redo_action.triggered.connect(self.undo_stack.redo)
        menu.addAction(redo_action)

        menu.exec(self.table_view.viewport().mapToGlobal(position))
        
    def eventFilter(self, obj, event):
        if obj == self.table_view and event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Tab:
                current_index = self.table_view.currentIndex()
                current_row = current_index.row()
                new_row = current_row + 1

                while new_row < self.table_model.rowCount() and self.table_view.isRowHidden(new_row):
                    new_row += 1

                if new_row < self.table_model.rowCount():
                    new_index = self.table_model.index(new_row, current_index.column())
                    self.table_view.setCurrentIndex(new_index)
                    return True  # Event handled
        return super().eventFilter(obj, event)
    
    def search_selected_cell(self):
        selected_index = self.table_view.currentIndex()
        if selected_index.isValid():
            search_text = str(self.table_model.data(selected_index, Qt.ItemDataRole.DisplayRole))
            self.find_field.setText(search_text)  #Populate the search field
            self.search(search_text) # Run the search
        else:
            QMessageBox.warning(self, "Upozorenje", "Niste odabrali ćeliju.")

    def on_data_changed(self, topLeft, bottomRight, roles):
        """Handles data changes in the table view."""
        self.unsaved_changes = True  # Mark unsaved changes
        self.status_bar.showMessage("Postoje nespremljene promjene.", 2000)
        self.table_view.viewport().update()  # Ensure visual updated
    
    def save_to_excel(self, file_path):
        df = self.table_model.get_data_frame()
        df.to_excel(file_path, index=False)

    def closeEvent(self, event):
        if self.table_model.has_unsaved_changes():
            reply = QMessageBox.question(
                self,
                "Nespremljene promjene",
                "Imate nespremljene promjene. Želite li ih spremiti prije izlaska?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.save_changes()
                event.accept()
            elif reply == QMessageBox.StandardButton.No:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()
